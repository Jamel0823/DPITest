import sys
import os
import pandas as pd
import openpyxl
import xlwt
import xlrd
import xlutils.copy
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment, colors
str = os.path.abspath(__file__)
str = str[0:-8]

def getTestCaceFromExcel(file, sheet):
    wb = openpyxl.load_workbook("%s"%file)
    ws = wb["%s"%sheet]
    rows, cols= ws.max_row, ws.max_column
    n = -1
    testCase = []
    for i in range(1, cols+1):
        tempItem = []
        ce = ws.cell(row=3, column=i)
        if ce.value == "V":
            n+=1
            tempItem.append(ws.cell(row=1, column=i).value)
            tempItem.append(ws.cell(row=2, column=i).value)
            for j in range(1,rows):
                if ws.cell(row=3+j, column=i).value != None:
                    tempItem.append(ws.cell(row=3+j, column=i).value)
            testCase.append(tempItem)
    wb.close()
    return testCase

def getParaFromExcel(file, testCase):
    para = []
    for i in range(0, len(testCase)):
        for j in range(0, len(testCase[i])):
            wb = openpyxl.load_workbook("%s"%file)
            ws = wb["%s_%s"%(testCase[i][1], testCase[i][0])]
            # print(testCase[i][0])
            rows, cols= ws.max_row, ws.max_column
            # print(testCase[i][j])
            if testCase[i][j] != "%s_%s_All"%(testCase[i][1], testCase[i][0]):
                tempPara = []
                for k in range(2, rows+1):
                    ce = ws.cell(row=k, column=1)
                    if ce.value == "%s"%testCase[i][j]:
                        # print(ce.value, testCase[i][j])
                        if len(tempPara) == 0:
                            # print(ws.cell(row=k, column=2).value)
                            tempPara.append(ws.cell(row=k, column=1).value)
                        tempPara.append(ws.cell(row=k, column=2).value)
                        # print(ws.cell(row=k, column=2).value)
                        # print(tempPara)
                        tempPara.append(ws.cell(row=k, column=3).value)
                        # print(tempPara)
                if len(tempPara) != 0:
                    para.append(tempPara)
                    # print(para)
            else:
                # print(case)
                case = ws.cell(row=2, column=2).value
                # print(case)
                tempPara = []
                for k in range(2, rows+1):
                    ce = ws.cell(row=k+1, column=2)
                    if ce.value == case:
                        # print(ce.value, case)
                        if len(tempPara) == 0:
                            tempPara.append("%s_%s_All"%(testCase[i][1], testCase[i][0]))
                        tempPara.append(ws.cell(row=k, column=1).value)
                        tempPara.append(ws.cell(row=k, column=2).value)
                        tempPara.append(ws.cell(row=k, column=3).value)
                    else:
                        # print(ce.value, case)
                        if len(tempPara) == 0:
                            tempPara.append("%s_%s_All"%(testCase[i][1], testCase[i][0]))
                        tempPara.append(ws.cell(row=k, column=1).value)
                        tempPara.append(ws.cell(row=k, column=2).value)
                        tempPara.append(ws.cell(row=k, column=3).value)
                        case = ce.value
                para.append(tempPara)
                # print(para)
            wb.close()
    # print(para)
    return para

def appendCasetoTXT(testCase):
    fp = open("%s\\Src\\Input\\Run.txt"%str, "w")
    # fp.write("%s"%testCase)
    for i in range(0, len(testCase)):
        fp.write("{")
        for j in range(0, len(testCase[i])):
            if j !=0 :
                fp.write(" ")
            fp.write("%s"%testCase[i][j])
        fp.write("} ")
    fp.close()
def appendParatoTXT(testCase, testPara):
    for i in range(0, len(testCase)):
        for j in range(2, len(testCase[i])):
            # print(testCase[i][j])
            fp = open("%s\\Src\\Input\\%s.txt"%(str, testCase[i][j]), "w")
            for m in range(0, len(testPara)):
                if testPara[m][0] == testCase[i][j]:
                    for n in range(1, len(testPara[m])):
                        fp.write("%s\n" % testPara[m][n])


testCase = getTestCaceFromExcel("%s\\Run.xlsm"%str, "TestFunction")
print(testCase)
testPara = getParaFromExcel("%s\\Run.xlsm"%str, testCase)
# print(testPara)
appendCasetoTXT(testCase)
appendParatoTXT(testCase, testPara)


# str = os.path.abspath(__file__)

# os.chdir(str + "\Src\Lib\Public\Excel")
# os.system("tclsh .\getInfor.tcl")
os.system(r'"C:\Program Files\Triangle MicroWorks\Protocol Test Harness\bin\tmwtest.exe"')
